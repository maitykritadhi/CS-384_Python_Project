from django.shortcuts import render
from django.http import JsonResponse
import openpyxl
from openpyxl.styles import Border, Side
import pandas as pd
import os
import shutil
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

def home(request):

    if request.method == 'POST':
        master_roll = request.FILES["master_roll"]
        response = request.FILES["response"]
        try:
            positive = int(request.POST["pmarks"])
            negative = int(request.POST["nmarks"])
        except:
            positive = float(request.POST["pmarks"])
            negative = float(request.POST["nmarks"])

        if 'one' == request.POST["action"]:

            # dir_output = 'output'
            # path_o = os.path.join('./',dir_output)

            # try:
            #     shutil.rmtree(path_o)
            # except OSError as e:
            #     pass

            # os.mkdir(path_o)

            master_file = master_roll.read().decode('utf-8')
            master_data = io.StringIO(master_file)
            master_df = pd.read_csv(master_data, sep=",")
            #print(master_df)

            response_file = response.read().decode('utf-8')
            response_data = io.StringIO(response_file)
            response_df = pd.read_csv(response_data, sep=",")
            #print(response_df)

            dict = {}
            for i in range(len(response_df)):            
                d = {}
                d['Timestamp'] = response_df.iloc[i,0]
                d['Email address'] = response_df.iloc[i,1]
                d['Score'] = response_df.iloc[i,2]
                d['Name'] = response_df.iloc[i,3]
                d['IITP webmail'] = response_df.iloc[i,4]
                d['Phone (10 digit only)'] = response_df.iloc[i,5]
                d['Roll Number'] = response_df.iloc[i,6]
                d['Answer'] = list(response_df.iloc[i,7:])

                dict[response_df.iloc[i,6]] = d

            if 'ANSWER' not in dict.keys():
                return JsonResponse({'foo':'bar'})

            ans_key = dict['ANSWER']['Answer']

            for x,y in dict.items():
                correct = 0
                incorrect = 0
                na = 0
                for i in range (len(ans_key)):
                    if(str(y['Answer'][i]) == 'nan'):
                        na += 1
                    elif(y['Answer'][i] == ans_key[i]):
                        correct += 1
                    else:
                        incorrect += 1

                dict[x]['correct'] = correct
                dict[x]['incorrect'] = incorrect
                dict[x]['na'] = na

            concise_df = response_df
            concise_df.insert(6,'Score_After_Negative','any')
            concise_df.insert(36, 'statusAns','any')

            for i in range(len(concise_df)):
                concise_df.iloc[i,6] = str(round(dict[response_df.iloc[i,7]]['correct']*positive + dict[response_df.iloc[i,7]]['incorrect']*negative, 2)) + '/' + str(round(positive*len(ans_key), 2))
                concise_df.iloc[i,36] = '[' + str(dict[response_df.iloc[i,7]]['correct']) + ',' + str(dict[response_df.iloc[i,7]]['incorrect']) + ',' + str(dict[response_df.iloc[i,7]]['na']) + ']'

            absname = []
            absroll = []
            absscore = []
            for i in range(len(master_df)):
                if master_df.iloc[i,0] not in dict:
                    absname.append(master_df.iloc[i,1])
                    absroll.append(master_df.iloc[i,0])
                    absscore.append('Absent')
            
            dabs = {'Roll Number':absroll, 'Name':absname, 'Score':absscore, 'Score_After_Negative':absscore}
            dfabs = pd.DataFrame(dabs)

            concise_df = concise_df.append(dfabs, ignore_index = True)

            concise_df.rename(columns={'Score': 'Google_Score'}, inplace=True)
            concise_df.to_csv('marksheets/'+'concise_marksheet.csv', index = False)


        if 'two' == request.POST["action"]:

            # dir_output = 'output'
            # path_o = os.path.join('./',dir_output)

            # try:
            #     shutil.rmtree(path_o)
            # except OSError as e:
            #     pass

            # os.mkdir(path_o)

            master_file = master_roll.read().decode('utf-8')
            master_data = io.StringIO(master_file)
            master_df = pd.read_csv(master_data, sep=",")
            #print(master_df)

            response_file = response.read().decode('utf-8')
            response_data = io.StringIO(response_file)
            response_df = pd.read_csv(response_data, sep=",")
            #print(response_df)

            dict = {}
            for i in range(len(response_df)):            
                d = {}
                d['Timestamp'] = response_df.iloc[i,0]
                d['Email address'] = response_df.iloc[i,1]
                d['Score'] = response_df.iloc[i,2]
                d['Name'] = response_df.iloc[i,3]
                d['IITP webmail'] = response_df.iloc[i,4]
                d['Phone (10 digit only)'] = response_df.iloc[i,5]
                d['Roll Number'] = response_df.iloc[i,6]
                d['Answer'] = list(response_df.iloc[i,7:])

                dict[response_df.iloc[i,6]] = d

            if 'ANSWER' not in dict.keys():
                return JsonResponse({'foo':'bar'})

            ans_key = dict['ANSWER']['Answer']

            for i in range(len(master_df)):
                if master_df.iloc[i,0] not in dict:
                    d = {}
                    d['Name'] = master_df.iloc[i,1]
                    d['Roll Number'] = master_df.iloc[i,0]
                    d['Answer'] = [float("nan")]*len(ans_key)
            
                    dict[master_df.iloc[i,0]] = d

            for x,y in dict.items():
                correct = 0
                incorrect = 0
                na = 0
                for i in range (len(ans_key)):
                    if(str(y['Answer'][i]) == 'nan'):
                        na += 1
                    elif(y['Answer'][i] == ans_key[i]):
                        correct += 1
                    else:
                        incorrect += 1

                dict[x]['correct'] = correct
                dict[x]['incorrect'] = incorrect
                dict[x]['na'] = na

            border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
            
            for x,y in dict.items():

                wb = openpyxl.Workbook()
                wb.create_sheet(index = 0, title = 'quiz')
                sheet = wb['quiz']
                img = openpyxl.drawing.image.Image('logo.jpeg')
                img.anchor = 'A1'
                sheet.add_image(img)

                sheet.column_dimensions['A'].width = 16.89
                sheet.column_dimensions['B'].width = 16.89
                sheet.column_dimensions['C'].width = 16.89
                sheet.column_dimensions['D'].width = 16.89
                sheet.column_dimensions['E'].width = 16.89

                sheet.row_dimensions[5].height = 22.8
                sheet. merge_cells('A5:E5')
                sheet['A5'].font = openpyxl.styles.Font(name = 'Century',size=18, bold=True,underline='single')
                sheet['A5'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                sheet['A5'] = 'Mark Sheet'

                #...............................................................

                sheet['A6'] = 'Name:'
                sheet['A6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
                sheet['A6'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='bottom')

                sheet['B6'] = y['Name']
                sheet['B6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['B6'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='bottom')

                sheet['D6'] = 'Exam:'
                sheet['D6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
                sheet['D6'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='bottom')

                sheet['E6'] = 'quiz'
                sheet['E6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['E6'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='bottom')

                sheet['A7'] = 'Roll Number:'
                sheet['A7'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
                sheet['A7'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='bottom')

                sheet['B7'] = y['Roll Number']
                sheet['B7'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['B7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='bottom')

                sheet['B9'] = 'Right'
                sheet['B9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['B9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['C9'] = 'Wrong'
                sheet['C9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['C9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['D9'] = 'Not Attempt'
                sheet['D9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['D9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['E9'] = 'Max'
                sheet['E9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['E9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['A10'] = 'No.'
                sheet['A10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['A10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['A11'] = 'Marking'
                sheet['A11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['A11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['A12'] = 'Total'
                sheet['A12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['A12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['B10'] = str(dict[x]['correct'])
                sheet['B10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
                sheet['B10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['B11'] = positive
                sheet['B11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
                sheet['B11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['B12'] = dict[x]['correct']*positive
                sheet['B12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
                sheet['B12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['C10'] = str(dict[x]['incorrect'])
                sheet['C10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
                sheet['C10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['C11'] = negative
                sheet['C11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
                sheet['C11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['C12'] = dict[x]['incorrect']*negative
                sheet['C12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
                sheet['C12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['D10'] = dict[x]['na']
                sheet['D10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
                sheet['D10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['D11'] = '0'
                sheet['D11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
                sheet['D11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['E10'] = str(len(ans_key))
                sheet['E10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
                sheet['E10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['E12'] =  str(round(dict[x]['correct']*positive + dict[x]['incorrect']*negative, 2)) + '/' + str(round(positive*len(ans_key), 2))
                sheet['E12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='0000FF')
                sheet['E12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheet['A15'] = 'Student Ans'
                sheet['A15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['A15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                sheet['A15'].border = border

                sheet['D15'] = 'Student Ans'
                sheet['D15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['D15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                sheet['D15'].border = border

                sheet['B15'] = 'Correct Ans'
                sheet['B15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['B15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                sheet['B15'].border = border

                sheet['E15'] = 'Correct Ans'
                sheet['E15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
                sheet['E15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                sheet['E15'].border = border

                #......................................................................

                for r in range(9,13):
                    for c in range(1,6):
                        sheet.cell(row = r , column = c).border = border

                for i in range(len(ans_key)):
                    if(i+16>40):
                        sheet['E'+str(i-9)].border = border
                        
                        sheet['E'+str(i-9)] = ans_key[i]
                        sheet['E'+str(i-9)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='0000FF')
                        sheet['E'+str(i-9)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                    else:   
                        sheet['B'+str(i+16)].border = border
                        
                        sheet['B'+str(i+16)] = ans_key[i]
                        sheet['B'+str(i+16)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='0000FF')
                        sheet['B'+str(i+16)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                        
                        
                for i in range(len(ans_key)):
                    if(i+16>40):
                        sheet['D'+str(i-9)].border = border
                        
                        if(str(dict[x]['Answer'][i]) == 'nan'):
                            pass
                        elif(dict[x]['Answer'][i] == ans_key[i]):
                            sheet['D'+str(i-9)] = dict[x]['Answer'][i]
                            sheet['D'+str(i-9)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
                            sheet['D'+str(i-9)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                        else:
                            sheet['D'+str(i-9)] = dict[x]['Answer'][i]
                            sheet['D'+str(i-9)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
                            sheet['D'+str(i-9)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')  
                        
                        
                    else:
                        sheet['A'+str(i+16)].border = border
                        
                        if(str(dict[x]['Answer'][i]) == 'nan'):
                            pass
                        elif(dict[x]['Answer'][i] == ans_key[i]):
                            sheet['A'+str(i+16)] = dict[x]['Answer'][i]
                            sheet['A'+str(i+16)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
                            sheet['A'+str(i+16)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                        else:
                            sheet['A'+str(i+16)] = dict[x]['Answer'][i]
                            sheet['A'+str(i+16)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
                            sheet['A'+str(i+16)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

                sheetDelete = wb["Sheet"]
                wb.remove(sheetDelete)
                wb.save("marksheets/"+x+'.xlsx')

        
        if 'three' == request.POST["action"]:

            master_file = master_roll.read().decode('utf-8')
            master_data = io.StringIO(master_file)
            master_df = pd.read_csv(master_data, sep=",")

            response_file = response.read().decode('utf-8')
            response_data = io.StringIO(response_file)
            response_df = pd.read_csv(response_data, sep=",")

            dict = {}
            for i in range(len(response_df)):            
                d = {}
                d['Timestamp'] = response_df.iloc[i,0]
                d['Email address'] = response_df.iloc[i,1]
                d['Score'] = response_df.iloc[i,2]
                d['Name'] = response_df.iloc[i,3]
                d['IITP webmail'] = response_df.iloc[i,4]
                d['Phone (10 digit only)'] = response_df.iloc[i,5]
                d['Roll Number'] = response_df.iloc[i,6]
                d['Answer'] = list(response_df.iloc[i,7:])

                dict[response_df.iloc[i,6]] = d


            fromaddr = "cs384python@gmail.com"

            # creates SMTP session
            s = smtplib.SMTP('smtp.gmail.com', 587)

            # start TLS for security
            s.starttls()

            # Authentication
            s.login(fromaddr, "Qwerty_1234")

            for x,y in dict.items():

                toaddr = dict[x]['Email address']
                file_path = "marksheets/" + x + ".xlsx"
                
                # instance of MIMEMultipart
                msg = MIMEMultipart()
                
                # storing the senders email address  
                msg['From'] = fromaddr
                
                # storing the receivers email address 
                msg['To'] = toaddr
                
                # storing the subject 
                msg['Subject'] = "CS384 Quiz marks"
                
                # string to store the body of the mail
                body = "Dear Students, \n\nCS384 Quiz marks are attached for referance. \n+" + str(positive) + " Correct, " + str(negative) + " for wrong"
                
                # attach the body with the msg instance
                msg.attach(MIMEText(body, 'plain'))
                
                # open the file to be sent 
                filename = file_path
                attachment = open(file_path, "rb")
                
                # instance of MIMEBase and named as p
                p = MIMEBase('application', 'octet-stream')
                
                # To change the payload into encoded form
                p.set_payload((attachment).read())
                
                # encode into base64
                encoders.encode_base64(p)
                
                p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
                
                # attach the instance 'p' to instance 'msg'
                msg.attach(p)
                
                # Converts the Multipart msg into a string
                text = msg.as_string()
                
                # sending the mail
                s.sendmail(fromaddr, toaddr, text)

            
            
        #......................................................................................

    if request.method == 'GET':
        dir_output = 'marksheets'
        path_o = os.path.join('./',dir_output)

        try:
            shutil.rmtree(path_o)
        except OSError as e:
            pass

        os.mkdir(path_o)

    return render(request, 'index.html')